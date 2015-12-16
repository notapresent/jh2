# -*- coding: utf-8 -*-
"""
    Record export functions
"""
from __future__ import unicode_literals
import datetime
import logging
import os
import csv
import itertools
import re
import glob

from sqlalchemy import func, or_, and_
import jinja2
from jinja2.filters import do_truncate
from openpyxl import Workbook
import xlwt

from ..models import Scan, Record, Genre, RecordFlag, Image, scan_records
from . import user_settings, export_manager
from rbm2m.action import genre_manager
from rbm2m.util import to_str

BATCH_SIZE = 10000
CSV_BATCH_SIZE = 500


MESHOK_CAT_MAP = {
    5: 13037,
    6: 2230,
    7: 2230,
    8: 2228,
    9: 2230,
    11: 2231,
    12: 2230,
    13: 13036,
    14: 13291,
    15: 13037,
    16: 13291,
    17: 2230,
    18: 2228,
    19: 13291,
    20: 13038,
    21: 2229,
    22: 13291,
    23: 13283,
    24: 13041,
    25: 13042,
    26: 13283,
    27: 13283,
    28: 13291,
    29: 13283,
    30: 2230,
    31: 2230,
    32: 13045,
    33: 13291,
    34: 13291,
    35: 13043,
    36: 13043
}

logger = logging.getLogger(__name__)


class Exporter(object):
    """
        Base class for exports. Builds record sets for export.
    """
    fmt = 'base'

    def __init__(self, session, filters=None):
        self.session = session
        self.filters = filters
        self.settings = user_settings.UserSettings(session)

    def log_export(self, ip, user_agent):
        """
            Save export entry and emit log message
        """
        expman = export_manager.ExportManager(self.session)
        expdata = {
            'user_agent': user_agent,
            'ip': ip,
            'format': self.fmt
        }
        exp = expman.from_dict(expdata)
        message = "{} export #{} for {}@{} started"
        logger.info(message.format(self.fmt, exp.id, user_agent, ip))
        return exp

    def latest_scans(self):
        """
            List of ids of last successful scans for each export-enabled genre
        """
        subquery = (
            self.session.query(Scan.id)
                .filter(Scan.status == 'success')
                .filter(Scan.genre_id == Genre.id)
                .order_by(Scan.started_at.desc())
                .limit(1)
                .as_scalar()
        )
        rows = (
            self.session.query(Genre.id, subquery)
                .filter(subquery.isnot(None))
                .filter(Genre.export_enabled.is_(True))
                .all()
        )
        return [scan_id for genre_id, scan_id in rows]

    def records(self, scan_ids):
        """
            Returns all records from scans in scan_ids, excluding the ones with
            'missing_images' and 'sold' status

            :param scan_ids: list of scan ids
            :return: generator producing Record values
        """
        batch_no = 0
        while True:
            query = (
                self.session.query(
                    scan_records.c.record_id.label('id'),
                    Record.artist, Record.title,
                    Record.label, Record.notes, Record.grade, Record.format,
                    Record.price, Record.genre_id,
                    Genre.title.label('genre_title'),
                    func.group_concat(Image.id, ' ').label('images'))
                .join(Record, Record.id == scan_records.c.record_id)
                .join(Genre, Genre.id == Record.genre_id)
                .outerjoin(Image, Image.record_id == scan_records.c.record_id)
                .outerjoin(RecordFlag,
                           RecordFlag.record_id == scan_records.c.record_id)
                .filter(scan_records.c.scan_id.in_(scan_ids))
                .filter(or_(
                    RecordFlag.name.is_(None),
                    ~RecordFlag.name.in_(['sold', 'missing_images'])))
                .order_by(scan_records.c.record_id)
                .group_by(scan_records.c.record_id)
            )

            if self.filters:
                for col_name, value in self.filters.items():
                    col = getattr(Record, col_name)
                    query = query.filter(col == value)

            records = query.offset(batch_no * BATCH_SIZE).limit(BATCH_SIZE).all()

            if not records:
                break

            for row in records:
                yield dict(zip(row.keys(), row))

            batch_no += 1

    def category_list(self):
        """
            List of exported categories
        """
        genman = genre_manager.GenreManager(self.session)
        return genman.exported_list()


class YMLExporter(Exporter):
    fmt = 'yml'

    def __init__(self, session, filters):
        super(YMLExporter, self).__init__(session, filters=filters)
        self.limit = self.export_limit()
        env = jinja2.Environment()
        template_string = self.settings['yml_description_template']['value']
        self.description_template = env.from_string(template_string)

    def export_limit(self):
        lim = self.settings['yml_export_limit']['value']
        return int(lim) if lim else False

    def generation_date(self):
        """
            Export set build date
        """
        return datetime.datetime.utcnow()

    def offers(self):
        """
            Generates a sequence of offers for YML export
        """
        scans = self.latest_scans()

        for num, rec in enumerate(self.records(scans)):
            if self.limit and num == self.limit:
                break

            yield self.make_offer(rec)

    def make_offer(self, rec):
        """
            Generates offer for YML from result row dictionary
        """
        offer = rec.copy()
        offer['title'] = format_title(rec['title'], rec['artist'])

        offer['description'] = self.format_description(rec)
        offer['price'] = self.make_yml_price(rec['price'])
        offer['images'] = self.format_yml_images(rec['images'])
        return offer

    def format_description(self, rec):
        """
            Format lot description accordint to template
        """
        return self.description_template.render(**rec)

    def make_yml_price(self, price):
        """
            Calculate lot price
        """
        formula = self.settings['formula_yml']['value']
        price = eval(formula, {'x': price})
        return int(round(price))

    def format_yml_images(self, image_ids):
        """
            Accepts list of space-separated image ids, returns list of image paths
        """
        rv = []
        if not image_ids:
            return rv

        if len(image_ids) > 8:  # No more than 8 images for each record in YML
            image_ids = image_ids[:8]

        for img_id in image_ids.replace(' ', '').split(','):
            rv.append(Image(id=img_id).make_filename('.jpg'))

        return rv


class TableExporter(Exporter):
    fmt = 'table'

    def rows(self):
        """
            Generates a sequence of rows for table export
        """
        scans = self.latest_scans()

        for rec in self.records(scans):
            yield self.make_row(rec)

    def make_row(self, rec):
        """
            Make table row from query result row
        """

        rec['price'] = self.make_price(rec['price'])
        return rec

    def make_price(self, price):
        """
            Calculate record price according to formula
        """
        formula = self.settings['formula_table']['value']
        return int(round(eval(formula, {'x': price})))


class XLSXExporter(TableExporter):

    def save(self, path):
        wb = Workbook(write_only=True, optimized_write=True)
        ws = wb.create_sheet()

        header = [
            'Артикул', 'Жанр', 'Формат',
            'Исполнитель', 'Название', 'Лейбл',
            'Состояние', 'Цена', 'Примечания'
        ]
        ws.append(header)

        for row in self.rows():
            ws.append([
                row['id'], row['genre_title'], row['format'],
                row['artist'], row['title'], row['label'],
                row['grade'], row['price'], row['notes']
            ])

        wb.save(path)


class XLSExporter(TableExporter):

    def __init__(self, session, filters):
        super(XLSExporter, self).__init__(session, filters)
        self.workbook = xlwt.Workbook()  # encoding='utf-8'

    def add_sheet(self, title):
        """
            Add new sheet with a header row to workbook
        """
        title = title.replace('/', ' and ')
        sheet = self.workbook.add_sheet(title)
        self.write_row(sheet, 0, [
            'Артикул',
            'Исполнитель', 'Название', 'Лейбл',
            'Состояние', 'Цена', 'Примечания'
        ])
        return sheet

    def save(self, path):
        sheets = {}  # genre id: sheet
        rowcounts = {}

        for row in self.rows():
            g_id, g_title = row['genre_id'], row['genre_title']

            if g_id in sheets:
                worksheet = sheets[g_id]
            else:
                worksheet = sheets[g_id] = self.add_sheet(g_title)

            current_row = rowcounts.setdefault(g_id, 1)

            self.write_row(worksheet, current_row, [
                row['id'],
                row['artist'], row['title'], row['label'],
                row['grade'], row['price'], row['notes']
            ])
            rowcounts[g_id] = current_row + 1

        if sheets:
            self.workbook.save(path)

    def write_row(self, sheet, row_no, values):
        """
            Write row to XLS worksheet
        """
        for idx, val in enumerate(values):
            sheet.write(row_no, idx, label=val)


class CSVExporter(TableExporter):
    HEADER = [
        'Тип продажи',
        'Наименование',
        'Код отдела',
        'Состояние товара',
        'Фиксированная цена',
        'Продолжительность',
        # 'Доставка по городу',
        # 'Доставка по стране',
        # 'Повторные торги',
        'Описание',
        'Метки',
        # 'Уведомления',
        'URL картинки'
    ]
    DESC_FMT = '<b>{} - {}</b><br>Label: {}<br>Состояние: {}<br>Дополнительно: {}, {}<br>#{}'

    def save(self, basepath):
        counter = 0
        chunk = list(itertools.islice(self.rows(), CSV_BATCH_SIZE))

        def make_row(rec):
            maxlength = 90 - 5 - len(rec['grade'])
            title = '{} - {}'.format(rec['artist'], rec['title'])
            title = do_truncate(title, maxlength, killwords=True, end='…')
            tags = "{},{}".format(tagfilter(rec['artist']), tagfilter(rec['title']))
            row = [
                'F',
                '{} ({})'.format(title, rec['grade']),
                MESHOK_CAT_MAP.get(rec['genre_id']),
                'Y' if rec['grade'] == 'Still Sealed' else 'N',
                rec['price'],
                '30',
                # '0',
                # '250',
                # '0',
                self.DESC_FMT.format(rec['artist'], rec['title'], rec['label'], rec['grade'], rec['format'],
                                     rec['notes'], rec['id']),
                tags,
                # 'Y',
                cover_url(rec['image_id'])
            ]
            return row

        while chunk:
            filename = "{}-{}.csv".format(basepath, counter)
            self.save_file(filename, map(make_row, chunk))
            chunk = list(itertools.islice(rows, CSV_BATCH_SIZE))
            counter += 1
            print counter

    def delete_old_csvs(self, basepath):
        fns = glob.glob("{}-*.csv")
        for fn in fns:
            os.unlink(fn)

    def save_file(self, filename, rows):
        with open(filename, 'wb') as csvfile:
            writer = csv.writer(csvfile)
            # writer.writerow(map(to_str, self.HEADER))

            for row in rows:
                row_ = [to_str(f) for f in row]
                writer.writerow(row_)

        with open(filename, 'rb+') as filehandle:
            filehandle.seek(-2, os.SEEK_END)
            filehandle.truncate()

    def records(self, scan_ids):
        """
            Returns all records from scans in scan_ids, excluding the ones with
            'missing_images' and 'sold' status

            :param scan_ids: list of scan ids
            :return: generator producing Record values
        """
        batch_no = 0
        while True:
            query = (
                self.session.query(
                    scan_records.c.record_id.label('id'),
                    Record.artist, Record.title,
                    Record.label, Record.notes, Record.grade, Record.format,
                    Record.price, Record.genre_id,
                    Genre.title.label('genre_title'),
                    Image.id.label('image_id')
                )
                .join(Record, Record.id == scan_records.c.record_id)
                .join(Genre, Genre.id == Record.genre_id)
                .outerjoin(Image, and_(Image.record_id == scan_records.c.record_id,
                                       Image.is_cover.is_(True)))
                .outerjoin(RecordFlag,
                           RecordFlag.record_id == scan_records.c.record_id)
                .filter(scan_records.c.scan_id.in_(scan_ids))
                .filter(or_(
                    RecordFlag.name.is_(None),
                    ~RecordFlag.name.in_(['sold', 'missing_images'])))
                .order_by(scan_records.c.record_id)
                .group_by(scan_records.c.record_id)
            )

            if self.filters:
                for col_name, value in self.filters.items():
                    col = getattr(Record, col_name)
                    query = query.filter(col == value)

            records = query.offset(batch_no * BATCH_SIZE).limit(BATCH_SIZE).all()

            if not records:
                break

            for row in records:
                yield dict(zip(row.keys(), row))

            batch_no += 1


def format_title(title, artist, max_length=50):
    """
        Truncate title if it is longer than max_length
    """
    max_title_length = max_length - len(artist) - 2
    truncated = do_truncate(title, max_title_length, killwords=True, end='…')
    return truncated


def cover_url(img_id):
    """
        Return small cover path
    """
    if not img_id:
        return None
    path = Image(id=img_id).make_filename('_small.jpg')
    baseurl = os.environ.get('MEDIA_BASEURL')
    return "{}/{}".format(baseurl, path)


def tagfilter(s, rx=re.compile(r'[^a-z0-9 ]', re.I)):
    """
        replace all non-alphanumeric characters
    """
    replace_map = {
        '&': ' and ',
        '\'': '',
    }

    def translate(match):
        return replace_map.get(match.group(0), ' ')

    s = rx.sub(translate, s)
    return re.sub(r'\s+', ' ', s).strip()
